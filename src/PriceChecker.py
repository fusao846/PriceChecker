import os
import sys
import re
import time
import json
import shutil
import threading
import random
import traceback
from datetime import datetime,timedelta
from pprint import pprint
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import TimeoutException

from common.Grids import CustomGrid, GRIDS as gr
from common.Chrome import Scraper
#from common.Chrome import Proxy
from common.Log import Log

import ScrapeLOEWE
import ScrapeLUISVUITTON
import ScrapeGUCCI
import ScrapeMIUMIU
import ScrapeBALENCIAGA
import ScrapePRADA
import ScrapeGeneral

VERSION = "v1.22"
# 0.95  2025/01/27 MIUMIU対応　Excelフォント設定　ファイル分割　スレッド制御　再開機能　URLチェック
# 0.96  2025/01/28 BALENCIAGA対応　　MIUMIUバグ修正
# 0.97  2025/01/28 PRADA対応
# 0.98  2025/01/28 BERLUTI対応　汎用スクレイパー対応
# 0.981 2025/01/29 サイズ改行、先頭０除去
# 0.99  2025/01/29 MONCLER対応
# 1.00  2025/01/30 完成版
# 1.01  2025/01/31 UKバグ修正
# 1.02  2025/01/31 SIZE 00 対応
# 1.03  2025/02/01 Moncler 英語ボタン対応
# 1.04  2025/02/02 LOEWE Size 1/2 対応 One size対応
# 1.05  2025/02/02 LV size 5.0 -> 5 
# 1.06  2025/02/02 LV Size 在庫なし対応 
# 1.07  2025/02/05 GUCCI 在庫なし不具合対応 
# 1.08  2025/02/11 BALENCIAGA サイズ　フランス⇒フランス/ヨーロッパ　対応 
# 1.09  2025/02/13 LUIS VUITTON サイズ仕様変更対応（有償） 
# 1.10  2025/02/16 LUIS VUITTON サイズ仕様変更対応 エラー対応 
# 1.11  2025/02/17 LUIS VUITTON サイズ仕様変更対応 クラス取得不備対応 
# 1.12  2025/02/18 LUIS VUITTON サイズ仕様変更対応 クラス取得不備対応 待機秒数調整
# 1.13  2025/02/18 LUIS VUITTON 魔改造
# 1.14  2025/02/18 LUIS VUITTON 待機時間調整
# 1.15  2025/02/23 LUIS VUITTON エラー処理追加
# 1.151 2025/02/23 LUIS VUITTON 2/20バージョンに戻す
# 1.16  2025/04/01 MARGIELA 新規追加
# 1.17  2025/04/06 LOEWE仕様変更対応
# 1.18  2025/04/24 Chrome With拡張対応
# 1.19  2025/04/25 サイズ例外対応
# 1.20  2025/05/03 NGワード対応
# 1.21  2025/06/29 GUCCIのサイズ　=NNcm 対応 
# 1.22  2025/07/09 Acceptボタン例外処理の修正

def strtobool(b):
    if b.upper() == "TRUE":
        return True
    else:
        return False

def priceNumber(text, euro=False):
    if euro == True:
        text = text.replace('.','')
        text = text.replace(',','.')
       # print(f"Euro ver {text}")
    cleaned_text = re.sub(r"[^0-9.]", "", text)  # 数字以外の文字を削除
    print(f"cleaned_test [{cleaned_text}]")
    number = int(float(cleaned_text))  # 整数に変換
    return number

def remove_zeros(s):
    return re.sub(r'^0+(\d)', r'\1', s)

def concat_size(textEles):
    size = ""
    otherSizeCount = 0;
    cnt = 0
    for tx in textEles:
        t = tx.get_attribute("textContent").strip()
        print(f"TTTTTT1 {t}")
        if "通知" in t or "サイズを選ぶ" in t or "Notify" in t or t == "サイズ" or "在庫なし" in t:
            None
        else:
            t = t.replace(",", ".")
            tSplit = t.split(" ")
            if len(tSplit) > 1 and  "フランス" in tSplit[0]:
                t = tSplit[1].strip()
            else:
                t = tSplit[0].strip()
                if len(tSplit) > 1:
                    if tSplit[1] == "1/2":
                        t = t + ".5"
                    if tSplit[0] == "One" and tSplit[1] == "size":
                        t = "OneSize"
            t = re.sub(r'=\d+(?:\.\d+)?cm', '', t, flags=re.IGNORECASE)
            t = t.replace("IT","")
            t = t.replace("UK","")
            t = t.replace(".0", "")
            print(f"TTTTTT2 {t}")
            if t == "0" or t == "00":
                None
            else:
                t = remove_zeros(t)
            print(f"[{t}]")
            if t in ["ミディアム" ,"ラージ","エクストララージ"]:
                otherSizeCount = otherSizeCount+1
            pattern = r'^[SLMX0-9.]+$'
            validSize = re.fullmatch(pattern, t)
            if size != "" and t != "" and validSize:
                size = size + "_"
            
            if validSize:
                size = size + t
    noZaiko = False
    if size == "" and otherSizeCount == 0:
        noZaiko = True
    if size == "" and otherSizeCount > 0:
        size = "OneSize"
    return size, noZaiko

CONFIG = {}
current = os.getcwd()
scraper = None
CHROME_FOLDER = "Chrome3"

STOP_FLAG = threading.Event()
thread = None

pcheck = f"{os.getenv('APPDATA')}\\PriceChecker"
json_file = f"{current}\\config.json"
try:
    with open(json_file, encoding='utf-8') as s:
        CONFIG = json.load(s)
except OSError as e:
    print(e)
    sys.exit()
LOG = Log("\\log\\PriceChecker3", strtobool(CONFIG['debug_mode']))
if getattr(sys, 'frozen', False):
    sys.stdout = open("log\\stdout.log", "w", encoding="utf-8")
    sys.stderr = open("log\\stderr.log", "w", encoding="utf-8")

LOG.put(f"Start version={VERSION}")
LOG.debug('init')
A=0
B=1
C=2
D=3
E=4
F=5
G=6
H=7
I=8
J=9
K=10
L=11
M=12

DEBUGGING=False
DEB = False
args = sys.argv
if len(args) > 1:
    DEB = True

def main():
    #print('★main')
    global CONFIG
    global LOG
    global DEB
    global thread

    def OKClick(cg):
        # スクレイピングを別スレッドで開始
        global thread
        STOP_FLAG.clear()
        thread = threading.Thread(target=OKClickMain, args=(cg,))
        thread.daemon = True  # メインスレッド終了時に自動終了
        thread.start()

    def OKClickMain(cg):
        global scraper
        #print('★OKClick')
        cg.setMessage("message", "")
        LOG.debug("OKClick")

        try:
            shutil.rmtree(f"{pcheck}\\{CHROME_FOLDER}")
        except Exception as e:
            None

        excel_path = cg.getValue("inExcel")
        workbook = None
        try:
            workbook = load_workbook(excel_path)
        except Exception as e:
            LOG.put(f"{e}")
            cg.setMessage("message", "定義Excelファイル名が無効です")
            return
        # シートを取得
        sheet = workbook.active  # 現在のアクティブなシート
        #urls = [cell.value for cell in sheet['L']]
        urls = []
        start_row = 0
        r = 2
        start_found = False
        for row in sheet.iter_rows(min_row=2,max_col=12,values_only=False):
            if row[H].value == None or row[H].value == "":
                if start_found == False:
                    start_found = True
                    start_row = r
            url = row[L].value
            if start_found == True and url:
                urls.append(url)
            if url == None or url == "":
                break
            r = r + 1
        if len(urls) == 0:
            cg.setMessage("message", "処理対象がありません")
            return
        try:
            print("scraper1")
            scraper = Scraper(f"{pcheck}\\{CHROME_FOLDER}", x=CONFIG["chrome_x"], y=CONFIG["chrome_y"])
        except Exception as e:
            LOG.put(f"{e}")
            cg.setMessage("message", f"{e}")
            return
        LOG.debug("scraper constructed 1")
        cnt = 0
        for r in range(0,len(urls)):
            cnt = cnt + 1
        cg.start_progress_bar("progress1", cnt)
        n = 0
        lv_renzoku = 0

        # メイン処理
        for row in sheet.iter_rows(min_row=start_row,max_col=12,values_only=False):
            if STOP_FLAG.is_set():
                return
            url = row[L].value
            if url == None or url == "":
                break
            n = n + 1
            price = 0
            size = ""
            zaiko = 0
            print('-----------------------------')
            # LOEWE
            if (url.startswith("https://www.loewe.com")):
                success = False
                for retry_count in range(CONFIG["retry_max"]):
                    if STOP_FLAG.is_set():
                        return

                    try:
                        price, size, zaiko = ScrapeGeneral.scrape("LOEWE", cg, scraper, url, LOG, concat_size, priceNumber)
                        success = True
                        break
                    except Exception as e:
                        LOG.put(f"retry count{retry_count} exception has occured {e}")
                        print(f"retry count{retry_count} exception has occured {e}")
                        traceback.print_exc() 
                        if scraper:
                            scraper.quit()
                        scraper = Scraper(f"{pcheck}\\Chrome2", x=CONFIG["chrome_x"], y=CONFIG["chrome_y"])
                if success == False:
                    cg.setMessage("message", "エラーが発生しました　中断します")
                    return
            # GUCCI
            if (url.startswith("https://www.gucci.com/jp")):
                success = False
                for retry_count in range(CONFIG["retry_max"]):
                    if STOP_FLAG.is_set():
                        return
                    try:
                        #price, size, zaiko = ScrapeGUCCI.scrapeGUCCI(cg, scraper, url, LOG, concat_size, priceNumber)
                        price, size, zaiko = ScrapeGeneral.scrape("GUCCI", cg, scraper, url, LOG, concat_size, priceNumber)
                        success = True
                        break
                    except Exception as e:
                        LOG.put(f"retry count{retry_count} exception has occured {e}")
                        print(f"retry count{retry_count} exception has occured {e}")
                        traceback.print_exc() 
                        if scraper:
                            scraper.quit()
                        scraper = Scraper(f"{pcheck}\\Chrome2", x=CONFIG["chrome_x"], y=CONFIG["chrome_y"])
                if success == False:
                    cg.setMessage("message", "エラーが発生しました　中断します")
                    return

            # BALENCIAGA
            if (url.startswith("https://www.balenciaga.com/")):
                success = False
                for retry_count in range(CONFIG["retry_max"]):
                    if STOP_FLAG.is_set():
                        return
                    try:
                        #price, size, zaiko = ScrapeBALENCIAGA.scrapeBALENCIAGA(cg, scraper, url, LOG, concat_size, priceNumber)
                        price, size, zaiko = ScrapeGeneral.scrape("BALENCIAGA", cg, scraper, url, LOG, concat_size, priceNumber)
                        success = True
                        break
                    except Exception as e:
                        LOG.put(f"retry count{retry_count} exception has occured {e}")
                        print(f"retry count{retry_count} exception has occured {e}")
                        traceback.print_exc() 
                        if scraper:
                            scraper.quit()
                        scraper = Scraper(f"{pcheck}\\Chrome2", x=CONFIG["chrome_x"], y=CONFIG["chrome_y"])
                if success == False:
                    cg.setMessage("message", "エラーが発生しました　中断します")
                    return

            # BERLUTI
            print(url)
            if (url.startswith("https://www.berluti.com/")):
                success = False
                for retry_count in range(CONFIG["retry_max"]):
                    if STOP_FLAG.is_set():
                        return
                    try:
                        price, size, zaiko = ScrapeGeneral.scrape("BERLUTI", cg, scraper, url, LOG, concat_size, priceNumber)
                        success = True
                        break
                    except Exception as e:
                        LOG.put(f"retry count{retry_count} exception has occured {e}")
                        print(f"retry count{retry_count} exception has occured {e}")
                        traceback.print_exc() 
                        if scraper:
                            scraper.quit()
                        scraper = Scraper(f"{pcheck}\\Chrome2", x=CONFIG["chrome_x"], y=CONFIG["chrome_y"])
                if success == False:
                    cg.setMessage("message", "エラーが発生しました　中断します")
                    return

            # MooRER
            if (url.startswith("https://www.moorer.clothing/")):
                success = False
                for retry_count in range(CONFIG["retry_max"]):
                    if STOP_FLAG.is_set():
                        return
                    try:
                        price, size, zaiko = ScrapeGeneral.scrape("MooRER", cg, scraper, url, LOG, concat_size, priceNumber)
                        success = True
                        break
                    except Exception as e:
                        LOG.put(f"retry count{retry_count} exception has occured {e}")
                        print(f"retry count{retry_count} exception has occured {e}")
                        traceback.print_exc() 
                        if scraper:
                            scraper.quit()
                        scraper = Scraper(f"{pcheck}\\Chrome2", x=CONFIG["chrome_x"], y=CONFIG["chrome_y"])
                if success == False:
                    cg.setMessage("message", "エラーが発生しました　中断します")
                    return

            # MONCLER
            if (url.startswith("https://www.moncler.com")):
                success = False
                for retry_count in range(CONFIG["retry_max"]):
                    if STOP_FLAG.is_set():
                        return
                    try:
                        price, size, zaiko = ScrapeGeneral.scrape("MONCLER", cg, scraper, url, LOG, concat_size, priceNumber)
                        print(f"{price} {size} {zaiko}")
                        success = True
                        break
                    except Exception as e:
                        LOG.put(f"retry count{retry_count} exception has occured {e}")
                        print(f"retry count{retry_count} exception has occured {e}")
                        traceback.print_exc() 
                        if scraper:
                            scraper.quit()
                        scraper = Scraper(f"{pcheck}\\Chrome2", x=CONFIG["chrome_x"], y=CONFIG["chrome_y"])
                if success == False:
                    cg.setMessage("message", "エラーが発生しました　中断します")
                    return

            # MARGIELA
            if (url.startswith("https://www.maisonmargiela.com")):
                success = False
                for retry_count in range(CONFIG["retry_max"]):
                    if STOP_FLAG.is_set():
                        return
                    try:
                        price, size, zaiko = ScrapeGeneral.scrape("MARGIELA", cg, scraper, url, LOG, concat_size, priceNumber)
                        print(f"{price} {size} {zaiko}")
                        success = True
                        break
                    except Exception as e:
                        LOG.put(f"retry count{retry_count} exception has occured {e}")
                        print(f"retry count{retry_count} exception has occured {e}")
                        traceback.print_exc() 
                        if scraper:
                            scraper.quit()
                        scraper = Scraper(f"{pcheck}\\Chrome2", x=CONFIG["chrome_x"], y=CONFIG["chrome_y"])
                if success == False:
                    cg.setMessage("message", "エラーが発生しました　中断します")
                    return


            # PRADA
            if (url.startswith("https://www.prada.com/")):
                success = False
                for retry_count in range(CONFIG["retry_max"]):
                    if STOP_FLAG.is_set():
                        return
                    try:
                        price, size, zaiko = ScrapePRADA.scrapePRADA(cg, scraper, url, LOG, concat_size, priceNumber)
                        success = True
                        break
                    except Exception as e:
                        LOG.put(f"retry count{retry_count} exception has occured {e}")
                        print(f"retry count{retry_count} exception has occured {e}")
                        traceback.print_exc() 
                        if scraper:
                            scraper.quit()
                        scraper = Scraper(f"{pcheck}\\Chrome2", x=CONFIG["chrome_x"], y=CONFIG["chrome_y"])
                if success == False:
                    cg.setMessage("message", "エラーが発生しました　中断します")
                    return
            # MIUMIU
            if (url.startswith("https://www.miumiu.com/")):
                success = False
                for retry_count in range(CONFIG["retry_max"]):
                    if STOP_FLAG.is_set():
                        return
                    try:
                        price, size, zaiko = ScrapeMIUMIU.scrapeMIUMIU(cg, scraper, url, LOG, concat_size, priceNumber)
                        success = True
                        break
                    except Exception as e:
                        LOG.put(f"retry count{retry_count} exception has occured {e}")
                        print(f"retry count{retry_count} exception has occured {e}")
                        traceback.print_exc() 
                        if scraper:
                            scraper.quit()
                        scraper = Scraper(f"{pcheck}\\Chrome2", x=CONFIG["chrome_x"], y=CONFIG["chrome_y"])
                if success == False:
                    cg.setMessage("message", "エラーが発生しました　中断します")
                    return

            if (url.startswith("https://jp.louisvuitton.com")):
                lv_renzoku = lv_renzoku + 1
                if lv_renzoku > CONFIG["lv_max"]:
                    LOG.debug(f"LV連続許容回数{CONFIG['lv_max']}を超えたので{CONFIG['lv_sleep']}秒待機")
                    cg.setMessage("message", f"{CONFIG['lv_sleep']}秒待機中")
                    print(f"LV連続許容回数{CONFIG['lv_max']}を超えたので{CONFIG['lv_sleep']}秒待機")
                    scraper.quit()
                    time.sleep(CONFIG["lv_sleep"])
                    scraper = Scraper(f"{pcheck}\\Chrome2", x=CONFIG["chrome_x"], y=CONFIG["chrome_y"])
                    cg.setMessage("message", f"")
                    LOG.debug("再開")
                    lv_renzoku = 0
                    #time.sleep(5)    

                success = False
                for retry_count in range(CONFIG["retry_max"]):
                    if STOP_FLAG.is_set():
                        return
                    try:
                        price, size, zaiko, Err = ScrapeLUISVUITTON.scrapeLUISVUITTON(cg, scraper, url,LOG,CONFIG,DEBUGGING,DEB,concat_size,priceNumber,pcheck, CHROME_FOLDER)
                        if Err != "":
                            cg.setMessage("message", "エラーが発生しました　中断します")
                            return
                        success = True
                        break
                    except Exception as e:
                        LOG.put(f"retry count{retry_count} exception has occured {e}")
                        print(f"retry count{retry_count} exception has occured {e}")
                        traceback.print_exc() 
                        if scraper:
                            scraper.quit()
                        scraper = Scraper(f"{pcheck}\\Chrome2", x=CONFIG["chrome_x"], y=CONFIG["chrome_y"])
                if success == False:
                    cg.setMessage("message", "エラーが発生しました　中断します")
                    return

            row[H].font = Font(name=CONFIG["excel_font_name"],size=CONFIG["excel_font_size"])
            row[J].font = Font(name=CONFIG["excel_font_name"],size=CONFIG["excel_font_size"])
            row[K].font = Font(name=CONFIG["excel_font_name"],size=CONFIG["excel_font_size"])
            row[J].value = price
            if price == "NOTFOUND":
                row[J].font = Font(color="FF0000",name=CONFIG["excel_font_name"],size=CONFIG["excel_font_size"])
            else:
                row[J].font = Font(color="000000",name=CONFIG["excel_font_name"],size=CONFIG["excel_font_size"])
            if zaiko == 0:
                row[H].font = Font(color="FF0000",name=CONFIG["excel_font_name"],size=CONFIG["excel_font_size"])
                row[H].value = "SOLD"
            else:
                row[H].font = Font(color="000000",name=CONFIG["excel_font_name"],size=CONFIG["excel_font_size"])
                row[H].value = size
            row[H].alignment = Alignment(horizontal="left")

            row[K].value = zaiko
            cg.set_progress("progress1", n)
            cg.update_idletasks()
            try:
                workbook.save(excel_path)
            except Exception as e:
                cg.setMessage("message", "Excelを保存できません")
                break
        scraper.quit()
        scraper = None
        cg.setMessage("message", "終了しました")
    def CancelClick(cg):
        LOG.debug("cancel")

    def ExitClick(cg):
        global scraper
        global thread
        result = cg.confirm("終了確認", "終了してもよろしいですか")
        if result:
            if scraper:
                STOP_FLAG.set()
                print("scraper quitting")
                scraper.quit()
                print("scraper done")

            if thread:
                if thread.is_alive():
                    print("thread alive")
                else:
                    print("thread done")

            if thread and thread.is_alive():
                thread.join()
            print("join done")
            LOG.debug("exit")
            cg.update()
            cg.destroy()
            sys.exit()
    def GoNext(cg):
        global DEBUGGING
        DEBUGGING = False

    settings = {
        "title":"価格チェックツール",
#        "title":"TEST",
        "width":600,
        "height":390,
        "left":5,
        "top":800,
        "grids":[
            {
                "type":gr.LABEL,
                "caption":f"価格チェックツール  {VERSION}"
#                "caption":"TEST"
            },
            {
                "type":gr.FILE_OPEN,
                "name":"inExcel",
                "label":"定義Excel",
                "title":"定義Excelファイルを選択してください",
                "init":"",
                "file_type":[("定義Excel", "*.xlsx")]
            },
            {
                "type":gr.PROGRESS_BAR,
                "name":"progress1",
            },
            {
                "type":gr.BUTTONS,
                "buttons":[
                    {
                       "caption":"実行",
                        "callback":OKClick
                    },
                    {
                        "caption":"終了",
                        "callback":ExitClick
                    },
                ]
            },
            {
                "type":gr.MESSAGE,
                "name":"message",
                "lines":6,
            }
        ]
    }
    if DEB == True:
        settings["grids"][3]["buttons"].append(
            {
                "caption":"N",
                "callback":GoNext
            }
        )

    cg = CustomGrid(settings)

if __name__ == '__main__':

    main()
    